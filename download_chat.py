"""
мң нҠңлёҢ лқјмқҙлёҢ мұ„нҢ… лӢӨмҡҙлЎңлҚ”
- лЎңк·ёмқё л¶Ҳн•„мҡ” мҳҒмғҒ: requestsлЎң л°”лЎң мҲҳм§‘
- лЎңк·ёмқё н•„мҡ” мҳҒмғҒ (л©ӨлІ„мӢӯ л“ұ): Chrome м„ём…ҳмңјлЎң мҝ нӮӨ мқёмҰқ нӣ„ мҲҳм§‘
- кІ°кіј: мӢңк°„ / мң нҳ• / лӢүл„Өмһ„ / л©”мӢңм§Җ / кёҲм•Ў вҶ’ CSV + Excel
"""

import sys
import subprocess

def _install(packages: list[str]):
    missing = []
    for pkg in packages:
        module = pkg.split("==")[0].replace("-", "_")
        # webdriver_manager мҳҲмҷё мІҳлҰ¬
        if module == "webdriver_manager":
            module = "webdriver_manager"
        try:
            __import__(module)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"[мһҗлҸҷ м„Өм№ҳ] н•„мҡ”н•ң нҢЁнӮӨм§ҖлҘј м„Өм№ҳн•©лӢҲлӢӨ: {', '.join(missing)}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])
        print("[мһҗлҸҷ м„Өм№ҳ] мҷ„лЈҢ.\n")

_install(["selenium", "webdriver-manager", "requests", "openpyxl"])

import csv
import json
import re
import time
from pathlib import Path

import hashlib
import time as _time

import requests
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager


CHROME_PROFILE = r"C:\Users\novaj\AppData\Local\Google\Chrome\User Data"


# в”Җв”Җ лЎңк·ёмқё м—Ҷмқҙ м„ём…ҳ мғқм„ұ (кіөк°ң мҳҒмғҒ) в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ

def make_session_without_login(video_id: str):
    """лЎңк·ёмқё м—Ҷмқҙ YouTube нҺҳмқҙм§Җм—җм„ң ytInitialData м¶”м¶ң"""
    url = f"https://www.youtube.com/watch?v={video_id}"
    print(f"нҺҳмқҙм§Җ лЎңл“ң мӨ‘: {url}", flush=True)

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept-Language": "ko-KR,ko;q=0.9",
    })
    resp = session.get(url, timeout=15)
    resp.raise_for_status()
    page_source = resp.text

    m = re.search(r"var ytInitialData\s*=\s*(\{.+?\});\s*(?:var |window\[|</script)", page_source, re.DOTALL)
    if not m:
        raise RuntimeError("ytInitialDataлҘј м°ҫм§Җ лӘ»н–ҲмҠөлӢҲлӢӨ. лЎңк·ёмқёмқҙ н•„мҡ”н•ң мҳҒмғҒмқј мҲҳ мһҲмҠөлӢҲлӢӨ.")
    initial_data = json.loads(m.group(1))

    m2 = re.search(r'"INNERTUBE_API_KEY"\s*:\s*"([^"]+)"', page_source)
    api_key = m2.group(1) if m2 else "AIzaSyAO_FJ2SlqU8Q4STEHLGCilw_Y9_11qcW8"

    m3 = re.search(r'"INNERTUBE_CONTEXT_CLIENT_VERSION"\s*:\s*"([^"]+)"', page_source)
    client_ver = m3.group(1) if m3 else "2.20240101.00.00"

    session.headers.update({
        "Referer": url,
        "Origin": "https://www.youtube.com",
    })
    print("нҺҳмқҙм§Җ лЎңл”© мҷ„лЈҢ (лЎңк·ёмқё м—ҶмқҢ)\n", flush=True)
    return session, initial_data, api_key, client_ver


# в”Җв”Җ SeleniumмңјлЎң мқёмҰқлҗң м„ём…ҳ мғқм„ұ (лЎңк·ёмқё н•„мҡ” мҳҒмғҒ) в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ

def make_session_via_chrome(video_id: str) -> tuple[requests.Session, str]:
    """Chrome н”„лЎңн•„лЎң YouTube нҺҳмқҙм§Җ м—ҙкі  вҶ’ мҝ нӮӨ + ytInitialData м¶”м¶ң"""

    print("\n" + "="*50, flush=True)
    print("мһ мӢң нӣ„ Chrome м°Ҫмқҙ мһҗлҸҷмңјлЎң м—ҙлҰҪлӢҲлӢӨ.", flush=True)
    print("Chrome м°Ҫм—җм„ң YouTubeм—җ лЎңк·ёмқён•ҙмЈјм„ёмҡ”.", flush=True)
    print("="*50, flush=True)
    while input("мӨҖл№„лҗҳм…Ёмңјл©ҙ ok мһ…л Ҙ: ").strip().lower() != "ok":
        pass

    print("ChromeDriver м„Өм№ҳ нҷ•мқё мӨ‘...", flush=True)
    driver_path = ChromeDriverManager().install()
    print(f"ChromeDriver мӨҖл№„ мҷ„лЈҢ", flush=True)

    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    print("Chrome мӢӨн–ү мӨ‘...", flush=True)
    driver = webdriver.Chrome(service=Service(driver_path), options=options)
    print("Chrome мӢӨн–үлҗЁ", flush=True)

    try:
        url = f"https://www.youtube.com/watch?v={video_id}"
        # лЁјм Җ мң нҠңлёҢ лЎңк·ёмқё нҺҳмқҙм§ҖлЎң мқҙлҸҷ
        print("YouTube лЎңк·ёмқё нҺҳмқҙм§ҖлҘј м—ҪлӢҲлӢӨ...", flush=True)
        driver.get("https://www.youtube.com")
        print("\n" + "="*50, flush=True)
        print("Chrome м°Ҫм—җм„ң YouTubeм—җ лЎңк·ёмқён•ҙмЈјм„ёмҡ”.", flush=True)
        print("лЎңк·ёмқё мҷ„лЈҢ нӣ„ мқҙ н„°лҜёл„җ м°ҪмңјлЎң лҸҢм•„мҷҖ okлҘј мһ…л Ҙн•ҙмЈјм„ёмҡ”.", flush=True)
        print("="*50, flush=True)
        while input("лЎңк·ёмқё мҷ„лЈҢ нӣ„ ok мһ…л Ҙ: ").strip().lower() != "ok":
            print("лЎңк·ёмқёмқ„ мҷ„лЈҢн•ң л’Ө okлҘј мһ…л Ҙн•ҙмЈјм„ёмҡ”.", flush=True)

        print(f"\nнҺҳмқҙм§Җ лЎңл“ң мӨ‘: {url}", flush=True)
        driver.get(url)

        print("нҺҳмқҙм§Җ лЎңл”© лҢҖкё° мӨ‘...", flush=True)
        WebDriverWait(driver, 15).until(
            lambda d: "ytInitialData" in d.page_source
        )
        print("нҺҳмқҙм§Җ лЎңл”© мҷ„лЈҢ", flush=True)

        # ytInitialData м¶”м¶ң
        page_source = driver.page_source
        m = re.search(r"var ytInitialData\s*=\s*(\{.+?\});\s*(?:var |window\[|</script)", page_source, re.DOTALL)
        if not m:
            raise RuntimeError("ytInitialDataлҘј м°ҫм§Җ лӘ»н–ҲмҠөлӢҲлӢӨ.")
        initial_data = json.loads(m.group(1))

        # л””лІ„к·ё: ytInitialData м ҖмһҘ
        debug_path = Path(__file__).parent / "_debug_initial.json"
        debug_path.write_text(json.dumps(initial_data, ensure_ascii=False, indent=2)[:100000], encoding="utf-8")
        print(f"  [л””лІ„к·ё] ytInitialData м ҖмһҘ: {debug_path}", flush=True)

        # API key м¶”м¶ң
        m2 = re.search(r'"INNERTUBE_API_KEY"\s*:\s*"([^"]+)"', page_source)
        api_key = m2.group(1) if m2 else "AIzaSyAO_FJ2SlqU8Q4STEHLGCilw_Y9_11qcW8"

        m3 = re.search(r'"INNERTUBE_CONTEXT_CLIENT_VERSION"\s*:\s*"([^"]+)"', page_source)
        client_ver = m3.group(1) if m3 else "2.20240101.00.00"

        # Selenium мҝ нӮӨ вҶ’ requests.Session
        all_cookies = driver.get_cookies()
        sapisid = next((c["value"] for c in all_cookies if c["name"] == "SAPISID"), None)
        sapisid3 = next((c["value"] for c in all_cookies if c["name"] == "__Secure-3PAPISID"), None)
        logged_in = bool(sapisid or sapisid3)

        print(f"лЎңк·ёмқё мғҒнғң: {'вң“ лЎңк·ёмқёлҗЁ' if logged_in else 'вң— лЎңк·ём•„мӣғ'}", flush=True)
        if not logged_in:
            raise RuntimeError("Chromeм—җм„ң YouTube лЎңк·ёмқёмқҙ н•„мҡ”н•©лӢҲлӢӨ.")

        # SAPISIDHASH н—ӨлҚ” мғқм„ұ (YouTube API мқёмҰқм—җ н•„мҲҳ)
        sid = sapisid3 or sapisid
        ts = int(_time.time())
        hash_val = hashlib.sha1(f"{ts} {sid} https://www.youtube.com".encode()).hexdigest()
        auth_header = f"SAPISIDHASH {ts}_{hash_val}"
        print(f"мқёмҰқ н—ӨлҚ” мғқм„ұ мҷ„лЈҢ", flush=True)

        session = requests.Session()
        session.headers.update({
            "User-Agent": driver.execute_script("return navigator.userAgent"),
            "Accept-Language": "ko-KR,ko;q=0.9",
            "Referer": url,
            "Origin": "https://www.youtube.com",
            "Authorization": auth_header,
            "X-Origin": "https://www.youtube.com",
        })
        for cookie in all_cookies:
            session.cookies.set(cookie["name"], cookie["value"], domain=cookie.get("domain", ".youtube.com"))

        return session, initial_data, api_key, client_ver

    finally:
        driver.quit()
        print("Chrome м°Ҫ лӢ«нһҳ\n", flush=True)


# в”Җв”Җ мұ„нҢ… replay API в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ

def find_continuation_token(data: dict) -> str | None:
    # 1мҲңмң„: м•Ңл Өм§„ кІҪлЎң м§Ғм ‘ нғҗмғү
    def dig(obj, keys):
        for k in keys:
            if not isinstance(obj, dict):
                return None
            obj = obj.get(k)
            if obj is None:
                return None
        return obj

    known_paths = [
        ["contents","twoColumnWatchNextResults","conversationBar","liveChatRenderer","continuations",0,"reloadContinuationData","continuation"],
        ["contents","twoColumnWatchNextResults","conversationBar","liveChatRenderer","continuations",0,"liveChatReplayContinuationData","continuation"],
        ["contents","twoColumnWatchNextResults","conversationBar","liveChatRenderer","continuations",0,"invalidationContinuationData","continuation"],
    ]
    for path in known_paths:
        val = data
        for k in path:
            if isinstance(val, list):
                val = val[k] if k < len(val) else None
            elif isinstance(val, dict):
                val = val.get(k)
            else:
                val = None
            if val is None:
                break
        if isinstance(val, str) and len(val) > 20:
            return val

    # 2мҲңмң„: м „мІҙ JSONм—җм„ң livechat кҙҖл Ё continuation нӮӨмӣҢл“ң нғҗмғү
    text = json.dumps(data)
    for pattern in [
        r'"reloadContinuationData"\s*:\s*\{[^}]*"continuation"\s*:\s*"([A-Za-z0-9_\-]{30,})"',
        r'"liveChatReplayContinuationData"\s*:\s*\{[^}]*"continuation"\s*:\s*"([A-Za-z0-9_\-]{30,})"',
        r'"continuation"\s*:\s*"([A-Za-z0-9_\-]{30,})"',
    ]:
        m = re.search(pattern, text)
        if m:
            return m.group(1)
    return None


def fetch_chat_page(session: requests.Session, continuation: str, api_key: str, client_ver: str, player_offset_ms: int = 0) -> dict:
    url = f"https://www.youtube.com/youtubei/v1/live_chat/get_live_chat_replay?key={api_key}"
    payload = {
        "context": {
            "client": {
                "clientName": "WEB",
                "clientVersion": client_ver,
                "hl": "ko",
            }
        },
        "continuation": continuation,
    }
    if player_offset_ms > 0:
        payload["currentPlayerState"] = {"playerOffsetMs": str(player_offset_ms)}
    for attempt in range(10):
        try:
            resp = session.post(url, json=payload, timeout=15)
            if resp.status_code in (503, 429, 500):
                wait = 2 ** attempt  # 1, 2, 4, 8, 16 ...
                print(f"  м„ңлІ„ мҳӨлҘҳ({resp.status_code}) вҖ” {wait}мҙҲ нӣ„ мһ¬мӢңлҸ„...", flush=True)
                _time.sleep(wait)
                continue
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            if attempt == 9:
                raise
            _time.sleep(2)
    raise RuntimeError("мөңлҢҖ мһ¬мӢңлҸ„ нҡҹмҲҳ мҙҲкіј")


def get_runs_text(runs: list) -> str:
    parts = []
    for r in runs:
        if "text" in r:
            parts.append(r["text"])
        elif "emoji" in r:
            sc = r["emoji"].get("shortcuts")
            parts.append(sc[0] if sc else "")
    return "".join(parts)


def sec_to_timestamp(sec: float) -> str:
    total = int(sec)
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    return f"{h}:{m:02d}:{s:02d}" if h else f"{m}:{s:02d}"


def parse_actions(actions: list) -> list[dict]:
    results = []
    for action in actions:
        # replayChatItemAction к°җмӢёкё° мІҳлҰ¬
        inner = action.get("replayChatItemAction", {}).get("actions", [action])
        for a in inner:
            item = a.get("addChatItemAction", {}).get("item", {})
            for renderer_key, msg_type in [
                ("liveChatTextMessageRenderer",    "мұ„нҢ…"),
                ("liveChatPaidMessageRenderer",    "мҠҲнҚјмұ—"),
                ("liveChatMembershipItemRenderer", "л©ӨлІ„мӢӯ"),
            ]:
                r = item.get(renderer_key)
                if not r:
                    continue
                offset_ms = int(
                    r.get("videoOffsetTimeMsec")
                    or r.get("timestampUsec", 0) and 0  # usecмқҖ лі„лҸ„ мІҳлҰ¬
                    or 0
                )
                # replayChatItemActionм—җм„ң videoOffsetTimeMsec к°Җм ёмҳӨкё°
                if offset_ms == 0:
                    offset_ms = int(action.get("replayChatItemAction", {}).get("videoOffsetTimeMsec", 0) or 0)
                author = r.get("authorName", {}).get("simpleText", "")
                text   = get_runs_text(r.get("message", {}).get("runs", []))
                amount = r.get("purchaseAmountText", {}).get("simpleText", "")
                if not text and msg_type == "л©ӨлІ„мӢӯ":
                    text = get_runs_text(r.get("headerSubtext", {}).get("runs", []))
                results.append({
                    "мӢңк°„":    sec_to_timestamp(offset_ms / 1000),
                    "мң нҳ•":    msg_type,
                    "лӢүл„Өмһ„":  author,
                    "л©”мӢңм§Җ":  text,
                    "кёҲм•Ў":    amount,
                    "_sec":   offset_ms / 1000,
                })
    return results


def _checkpoint_path(output_dir: Path, base_name: str) -> Path:
    return output_dir / f"{base_name}_мІҙнҒ¬нҸ¬мқёнҠё.json"


def _save_checkpoint(messages: list[dict], output_dir: Path, base_name: str, continuation: str | None = None):
    fields = ["мӢңк°„", "мң нҳ•", "лӢүл„Өмһ„", "л©”мӢңм§Җ", "кёҲм•Ў"]
    path = output_dir / f"{base_name}_мұ„нҢ…м „мІҙ.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(messages)
    if continuation is not None:
        cp = _checkpoint_path(output_dir, base_name)
        cp.write_text(json.dumps({"continuation": continuation, "count": len(messages)}, ensure_ascii=False), encoding="utf-8")
    print(f"  [мӨ‘к°„м ҖмһҘ] {len(messages)}к°ң вҶ’ {path.name}", flush=True)


def _load_existing_csv(output_dir: Path, base_name: str) -> tuple[list[dict], set] | None:
    csv_path = output_dir / f"{base_name}_мұ„нҢ…м „мІҙ.csv"
    if not csv_path.exists():
        return None
    messages = []
    keys = set()
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            messages.append(row)
            keys.add((row["лӢүл„Өмһ„"], row["л©”мӢңм§Җ"], row["кёҲм•Ў"]))
    return messages, keys


def select_output_dir() -> Path:
    """GUI нҸҙлҚ” м„ нғқ лӢӨмқҙм–јлЎңк·ёлЎң м ҖмһҘ мң„м№ҳлҘј м„ нғқ. мӢӨнҢЁ мӢң м§Ғм ‘ мһ…л Ҙ."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        folder = filedialog.askdirectory(
            title="мұ„нҢ… кё°лЎқ м ҖмһҘ мң„м№ҳлҘј м„ нғқн•ҳм„ёмҡ”",
            initialdir=Path.home() / "Desktop",
        )
        root.destroy()
        if folder:
            return Path(folder)
    except Exception as e:
        print(f"нҸҙлҚ” м„ нғқ м°Ҫмқ„ м—ҙм§Җ лӘ»н–ҲмҠөлӢҲлӢӨ: {e}")

    print("м ҖмһҘн•  нҸҙлҚ” кІҪлЎңлҘј м§Ғм ‘ мһ…л Ҙн•ҳм„ёмҡ” (л№„мӣҢл‘җл©ҙ мҠӨнҒ¬лҰҪнҠё нҸҙлҚ”м—җ м ҖмһҘ):")
    path_input = input("кІҪлЎң: ").strip()
    if path_input:
        p = Path(path_input)
        p.mkdir(parents=True, exist_ok=True)
        return p
    return Path(__file__).parent


def fetch_all_chat(session: requests.Session, initial_data: dict, api_key: str, client_ver: str, video_id: str = "chat", output_dir: Path | None = None) -> list[dict]:
    if output_dir is None:
        output_dir = Path(__file__).parent
    cp_path = _checkpoint_path(output_dir, video_id)

    existing_messages: list[dict] = []
    existing_keys: set = set()
    fast_forward = False

    # л°©лІ• 1: мІҙнҒ¬нҸ¬мқёнҠё JSON + CSV вҶ’ нҶ нҒ° м§Ғм ‘ мқҙм–ҙл°ӣкё°
    if cp_path.exists():
        try:
            cp_data = json.loads(cp_path.read_text(encoding="utf-8"))
            continuation = cp_data["continuation"]
            loaded = _load_existing_csv(output_dir, video_id)
            if loaded:
                existing_messages, existing_keys = loaded
            print(f"[мІҙнҒ¬нҸ¬мқёнҠё мқҙм–ҙл°ӣкё°] {len(existing_messages)}к°ң вҶ’ м ҖмһҘлҗң нҶ нҒ°л¶Җн„° мһ¬к°ң", flush=True)
        except Exception as e:
            print(f"мІҙнҒ¬нҸ¬мқёнҠё мҶҗмғҒ({e}), CSV мқҙм–ҙл°ӣкё° мӢңлҸ„...", flush=True)
            cp_path.unlink(missing_ok=True)
            continuation = None

    # л°©лІ• 2: CSVл§Ң мһҲмқҢ вҶ’ л§Ҳм§Җл§ү л©”мӢңм§Җ лӮҙмҡ©мңјлЎң лҒҠкёҙ кіі нғҗмғү
    if not cp_path.exists():
        loaded = _load_existing_csv(output_dir, video_id)
        if loaded:
            existing_messages, existing_keys = loaded
            # л§Ҳм§Җл§ү 50к°ң л©”мӢңм§Җмқҳ (лӢүл„Өмһ„+л©”мӢңм§Җ) м§‘н•©
            tail_keys = set()
            for m in existing_messages[-50:]:
                tail_keys.add((m.get("лӢүл„Өмһ„",""), m.get("л©”мӢңм§Җ","")))
            print(f"[CSV мқҙм–ҙл°ӣкё°] кё°мЎҙ {len(existing_messages)}к°ң лЎңл“ң вҖ” лҒҠкёҙ кіі нғҗмғү мӨ‘...", flush=True)
            fast_forward = True
        continuation = find_continuation_token(initial_data)
        if not continuation:
            raise RuntimeError("мұ„нҢ… replay нҶ нҒ°мқ„ м°ҫм§Җ лӘ»н–ҲмҠөлӢҲлӢӨ.")
        if not existing_messages:
            print(f"[мғҲлЎң мҲҳм§‘]", flush=True)

    print("мұ„нҢ… мҲҳм§‘ мӨ‘...", flush=True)
    new_messages: list[dict] = []
    page = 0
    tail_matched = False
    seen_tokens: set[str] = set()  # л¬ҙн•ңл°ҳліө л°©м§Җ

    while continuation:
        page += 1
        if continuation in seen_tokens:
            print(f"  нҶ нҒ° мӨ‘ліө к°җм§Җ вҖ” мҲҳм§‘ мҷ„лЈҢлЎң нҢҗлӢЁн•ҳкі  мў…лЈҢн•©лӢҲлӢӨ.", flush=True)
            break
        seen_tokens.add(continuation)
        data = fetch_chat_page(session, continuation, api_key, client_ver)

        live_chat = (
            data.get("continuationContents", {}).get("liveChatContinuation")
            or data.get("contents", {}).get("liveChatContinuation")
            or {}
        )

        actions = live_chat.get("actions", [])
        parsed = parse_actions(actions)

        if fast_forward:
            # мқҙ нҺҳмқҙм§Җ л§Ҳм§Җл§ү л©”мӢңм§Җк°Җ tail_keysм—җ мһҲмңјл©ҙ м•„м§Ғ кө¬к°„ м•Ҳ вҶ’ нҶөм§ёлЎң мҠӨнӮө
            if parsed:
                last_msg = parsed[-1]
                if (last_msg["лӢүл„Өмһ„"], last_msg["л©”мӢңм§Җ"]) in tail_keys:
                    tail_matched = True
                    if page % 100 == 0:
                        print(f"  [нғҗмғүмӨ‘] {page}нҺҳмқҙм§Җ мҠӨнӮө...", flush=True)
                    # нҺҳмқҙм§Җ м „мІҙ мҠӨнӮө
                elif tail_matched:
                    # tail_keys кө¬к°„мқ„ л§ү л„ҳм–ҙм„¬ вҶ’ м—¬кё°м„ңл¶Җн„° мҲҳм§‘
                    fast_forward = False
                    print(f"  лҒҠкёҙ кіі л°ңкІ¬ (нҺҳмқҙм§Җ {page})! мҲҳм§‘ мһ¬к°ңн•©лӢҲлӢӨ.", flush=True)
                    new_on_page = [m for m in parsed
                                   if (m["лӢүл„Өмһ„"], m["л©”мӢңм§Җ"]) not in existing_keys]
                    new_messages.extend(new_on_page)
                else:
                    # tail_keysлҘј м•„м§Ғ лӘ» лҙ„ вҶ’ мҠӨнӮө кі„мҶҚ
                    if page % 100 == 0:
                        print(f"  [нғҗмғүмӨ‘] {page}нҺҳмқҙм§Җ...", flush=True)
        else:
            new_messages.extend(parsed)
            if page % 50 == 0:
                total = len(existing_messages) + len(new_messages)
                print(f"  {total}к°ң мҲҳм§‘ мӨ‘...", flush=True)

        continuation = None
        for c in live_chat.get("continuations", []):
            for key in c:
                token = c[key].get("continuation")
                if token:
                    continuation = token
                    break
            if continuation:
                break

        # 500нҺҳмқҙм§Җл§ҲлӢӨ мІҙнҒ¬нҸ¬мқёнҠё м ҖмһҘ (нғҗмғү мҷ„лЈҢ нӣ„м—җл§Ң)
        if not fast_forward and page % 500 == 0:
            all_so_far = existing_messages + new_messages
            _save_checkpoint(all_so_far, output_dir, video_id, continuation)

    # мҷ„лЈҢ мӢң мІҙнҒ¬нҸ¬мқёнҠё мӮӯм ң
    if cp_path.exists():
        cp_path.unlink()

    all_messages = existing_messages + new_messages
    print(f"мҙқ {len(all_messages)}к°ң (кё°мЎҙ {len(existing_messages)}к°ң + мғҲлЎң {len(new_messages)}к°ң)", flush=True)
    return all_messages


# в”Җв”Җ м ҖмһҘ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ

def save_results(messages: list[dict], output_dir: Path, base_name: str):
    fields = ["мӢңк°„", "мң нҳ•", "лӢүл„Өмһ„", "л©”мӢңм§Җ", "кёҲм•Ў"]

    csv_path = output_dir / f"{base_name}_мұ„нҢ…м „мІҙ.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(messages)
    print(f"[CSV]   {csv_path}")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "м „мІҙ мұ„нҢ…"
        ws.append(fields)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1565C0")
            cell.alignment = Alignment(horizontal="center")

        type_colors = {"мҠҲнҚјмұ—": "FFCDD2", "л©ӨлІ„мӢӯ": "C8E6C9", "мұ„нҢ…": "FFFFFF"}
        for m in messages:
            ws.append([m["мӢңк°„"], m["мң нҳ•"], m["лӢүл„Өмһ„"], m["л©”мӢңм§Җ"], m["кёҲм•Ў"]])
            color = type_colors.get(m["мң нҳ•"], "FFFFFF")
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor=color)

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 60
        ws.column_dimensions["E"].width = 14

        xlsx_path = output_dir / f"{base_name}_мұ„нҢ….xlsx"
        wb.save(xlsx_path)
        print(f"[Excel] {xlsx_path}")
    except Exception as e:
        print(f"Excel м ҖмһҘ мӢӨнҢЁ: {e}")


# в”Җв”Җ мӢӨн–ү в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ

def extract_video_id(url: str) -> str:
    m = re.search(r"(?:v=|youtu\.be/)([A-Za-z0-9_\-]{11})", url)
    return m.group(1) if m else url.strip()


def main():
    print("=== мң нҠңлёҢ лқјмқҙлёҢ мұ„нҢ… лӢӨмҡҙлЎңлҚ” ===\n")

    url = (sys.argv[1] if len(sys.argv) >= 2
           else input("мң нҠңлёҢ мҳҒмғҒ URL л¶ҷм—¬л„Јкё°: ").strip())
    if not url:
        input("URL м—ҶмқҢ. м—”н„°лЎң мў…лЈҢ.")
        sys.exit(0)

    video_id = extract_video_id(url)

    print("\nмұ„нҢ… кё°лЎқмқ„ м ҖмһҘн•  нҸҙлҚ”лҘј м„ нғқн•ҙмЈјм„ёмҡ”.")
    output_dir = select_output_dir()
    print(f"м ҖмһҘ мң„м№ҳ: {output_dir}\n")

    ans = input("кө¬кёҖ лЎңк·ёмқёмқҙ мһҲм–ҙм•ј м ‘к·ј к°ҖлҠҘн•ң мҳҒмғҒмқёк°Җмҡ”? (л©ӨлІ„мӢӯ л“ұ) [y/n]: ").strip().lower()
    need_login = ans == "y"

    try:
        if need_login:
            session, initial_data, api_key, client_ver = make_session_via_chrome(video_id)
        else:
            session, initial_data, api_key, client_ver = make_session_without_login(video_id)
        messages = fetch_all_chat(session, initial_data, api_key, client_ver, video_id, output_dir)
    except Exception as e:
        print(f"\n[мҳӨлҘҳ] {e}")
        input("\nм—”н„°лҘј лҲ„лҘҙл©ҙ м°Ҫмқҙ лӢ«нһҷлӢҲлӢӨ.")
        sys.exit(1)

    save_results(messages, output_dir, video_id)
    print(f"\nмҷ„лЈҢ! кІ°кіј нҢҢмқјмқҙ м ҖмһҘлҗҳм—ҲмҠөлӢҲлӢӨ.\nм ҖмһҘ мң„м№ҳ: {output_dir}")
    input("\nм—”н„°лҘј лҲ„лҘҙл©ҙ м°Ҫмқҙ лӢ«нһҷлӢҲлӢӨ.")


if __name__ == "__main__":
    main()
